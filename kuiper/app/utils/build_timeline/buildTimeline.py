
import openpyxl
from datetime import datetime
import yaml
import re
import random 
import warnings
import os 

#from pykwalify.core import Core
warnings.filterwarnings('ignore' , category=UserWarning, module="openpyxl")

# get the value from a json object based on provided path
def json_get_val_by_path(j , p):
	try:
		p_l = p.split('.')
		if len(p_l) == 1:
			return [True, j[p] ]
		k = p_l[0]
		if k not in j.keys():
			return [False, "Key ["+str(k)+"] not in the json"]
			
		return json_get_val_by_path( j[k] , '.'.join(p_l[1:]))
	except Exception as e:
		return [False , str(e)]
		
class BuildTimeline:

	number_of_rows 	= 1
	color_sheet		= "Colors"
	views_folder	= "./views/"

	def __init__(self, views_folder="./views" , fname = None):
		self.views_folder=views_folder
		# if the original file name not provided, then mostly this will only used for generic functions (get_views, set_views, etc)
		if fname is not None:
			self.fname = fname
			self.wb = openpyxl.load_workbook(fname)

	
	# this will generate a random style for the provided column,row 
	def generate_rand_style(self, sheet, row, column):

		r = random.randint(150,255)
		g = random.randint(150,255)
		b = random.randint(150,255)

		rand_color = "%02x%02x%02x" % (r,g,b)
		backgroundColor = openpyxl.styles.colors.Color(rgb=rand_color)
		backgoundFill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=backgroundColor)
		cell = sheet.cell(row=row , column=column)
		cell.fill = backgoundFill
		return cell._style

	# get style from the sheet based on the column
	def get_style_from_sheet(self, column, value):
		ws = self.get_sheet_by_name(self.color_sheet)
		headers = self.get_sheet_headers(self.color_sheet)
		auto_generate = None
		for row in ws.iter_rows(min_row=2):
			if column == row[0].value:

				if row[1].value == "equal" and value == row[2].value:
					return row[3]._style
				elif row[1].value == "startswith" and value.startswith(row[2].value):
					return row[3]._style
				elif row[3].value == "auto":
					auto_generate = row

		if auto_generate is not None:
			match = re.match( auto_generate[2].value , value)
			if match:
				new_style_row = [auto_generate[0].value , auto_generate[1].value , match.group() , "" ]				
				new_row_count = ws.max_row+1
				style = self.generate_rand_style(sheet=ws , row=new_row_count , column= 3)
				for i in range( 0 , len(new_style_row)):
					if i == 3:
						self.add_value_to_sheet(sheet_name=self.color_sheet , row=new_row_count , column=i+1 , value=new_style_row[i] , style=style)
					else:
						self.add_value_to_sheet(sheet_name=self.color_sheet , row=new_row_count , column=i+1 , value=new_style_row[i])
				return style
			return None

	# get sheet object 
	def get_sheet_by_name(self, name):
		return self.wb.get_sheet_by_name(name)

	# get sheet headers
	def get_sheet_headers(self, sheet_name):
		headers = []
		ws = self.get_sheet_by_name(sheet_name)
		for header in ws[1]:
			if header.value is not None:
				headers.append(header.value)
		return headers

	# insert a value to the sheet
	def add_value_to_sheet(self, sheet_name, row , column , value , style=None):
		ws = self.get_sheet_by_name(sheet_name)
		cell = ws.cell(row=row , column=column)
		cell.value = value
		cell._style= style

	# add data object to the sheet, data object contains information about the data added
	def add_data_to_sheet(self, sheet_name , data):
		headers = self.get_sheet_headers(sheet_name)
		ws = self.get_sheet_by_name(sheet_name)
		new_row_count = ws.max_row+1
		for d_header in data.keys():
			header_index = headers.index(d_header)+1

			style = self.get_style_from_sheet(column=d_header , value=data[d_header])
			self.add_value_to_sheet(sheet_name=sheet_name , row=new_row_count , column=header_index , value=data[d_header] , style=style)

	# save the generated timeline to a file
	def save(self, path):
		self.wb.save(path)
		return path
		 
	# map the fields configuration to the columns
	def merge_data_and_fields(self, fields, data):
		for f in fields.keys():
			res = re.subn(r"(\$\{[a-zA-Z0-9@\._]*\})" , lambda x: str(json_get_val_by_path(data, x.group().lstrip("${").rstrip("}"))[1])  , fields[f])
			fields[f] = res[0]

		return fields
	
	# get a list of values based on sheet and column
	def get_values_by_column(self, sheet_name, column):
		headers = self.get_sheet_headers(sheet_name)
		header_index = headers.index(column)+1
		res =[]
		ws = self.get_sheet_by_name(sheet_name)
		for cell in ws.iter_rows(header_index , ws.max_row):
			res.append( cell[0].value )
		return res

	# get the list of views from the views table
	def get_views(self, views_path=None):
		if views_path is None: 
			views_path = os.path.abspath(self.views_folder)
		all_views = []
		for r,d,f in os.walk(views_path): 
			for file in f: 
				file_handle = open( os.path.join(views_path , file) , 'r' )
				file_content = file_handle.read()
				views = yaml.load( file_content, Loader=yaml.FullLoader )
				for v in range(len(views)):
					views[v]['path']	= file
					views[v]['content'] = file_content

				all_views += views
			for directory in d:
				all_views += self.get_views(os.path.join(views_path , directory))
		return all_views

	# save a view to file
	def set_views(self, yaml_content, path):
		try:
			dest_path = os.path.abspath(path)
			
			with open(dest_path , 'w') as fout:
				fout.write(yaml_content)
				return (True , "File ["+dest_path+"] modified")
		except Exception as e:
			return (False , str(e))

	# delete view from views folder
	def delete_views(self, name , path):
		try:
			dest_path = os.path.abspath(path)

			views = yaml.load( open(dest_path , 'r'), Loader=yaml.FullLoader )
			for i in range(len(views)):
				if views[i]['name'] == name:
					views.pop(i)
			
			# if the file has one view and deleted, then delete the file
			if len(views) == 0: 
				os.remove(dest_path)
				return (True, "File ["+dest_path+"] deleted")

			yaml_from_json = yaml.dump(views)
			set_v = self.set_views(yaml_from_json , path)
			if set_v[0]:
				return (True , "View ["+name+"] deleted")
			else:
				return (False, "Failed to delete view: " + set_v[1])

		except Exception as e:
			return (False, str(e))
 
	
	# validate the view data
	def validate_view(self , content):
		try:
			y = yaml.load(content , Loader=yaml.FullLoader)
			if not isinstance(y , list):
				return (False, "The view should be a yaml list")

				 
			for i in y:
				important_keys = ["fields" , 'name' , 'sheet' , 'condition']
				for k in important_keys:
					if k not in i:
						return (False, '"'+k+'" field not specified in the yaml: ' + str(i) )

			return (True, 'success')
		except Exception as e:
			return (False, str(e))
